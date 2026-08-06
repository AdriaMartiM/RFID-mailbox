# app/main.py
import sys

# La consola de Windows usa cp1252 por defecto; forzamos UTF-8 para evitar
# errores al imprimir caracteres acentuados.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import os
import uuid
import unicodedata
from typing import Optional
from datetime import datetime, timedelta, time
from fastapi import FastAPI, Depends, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from pydantic import BaseModel
from app.database import SessionLocal, engine, Base
from app.services.ticket_service import TicketService
import app.models as models

import re

# Estados válidos del sistema (deben coincidir con los del front-end)
ESTADOS_VALIDOS = {"PROCESADO", "EN_PROGRESO", "CERRADO"}

# SLA: horas LABORABLES objetivo para dar la primera respuesta a una incidencia.
# Si una incidencia abierta supera este tiempo sin que respondamos, salta el aviso.
SLA_RESPUESTA_HORAS = 8


def _es_nuestro(remitente: str) -> bool:
    """¿El mensaje lo hemos enviado nosotros (el equipo)?"""
    return bool(re.search(r"sistema|software|stcretail|rfid@", remitente or "", re.I))


# Palabras vacías: artículos, preposiciones, pronombres, conjunciones, verbos y
# muletillas comunes. No aportan nada para comparar incidencia ↔ pregunta.
_STOPWORDS = {
    # artículos / determinantes
    "los", "las", "unos", "unas", "una", "uno", "del", "este", "esta", "esto", "estos",
    "estas", "ese", "esa", "eso", "esos", "esas", "aquel", "aquella", "aquello", "cada",
    "algun", "alguna", "algunos", "algunas", "otro", "otra", "otros", "otras", "mismo", "misma",
    # preposiciones / conjunciones
    "ante", "bajo", "con", "contra", "desde", "durante", "entre", "hacia", "hasta",
    "mediante", "para", "por", "segun", "sin", "sobre", "tras", "pero", "sino", "porque",
    "pues", "aunque", "como", "cuando", "donde", "mientras", "tambien", "ademas",
    # pronombres
    "que", "quien", "cual", "cuales", "nos", "les", "sus", "mis", "tus", "ella", "ellos",
    "ellas", "nosotros", "vosotros", "ustedes", "usted",
    # verbos / formas comunes
    "ser", "soy", "eres", "somos", "son", "fue", "fueron", "era", "eran", "estar", "estan",
    "estoy", "estamos", "estaba", "estaban", "esta", "tener", "tiene", "tienen", "tengo",
    "tenia", "teniendo", "hacer", "hace", "hacen", "haciendo", "puede", "pueden", "puedo",
    "poder", "podria", "seguir", "seguimos", "sigue", "siendo", "habia", "haya", "hemos",
    "ver", "viendo", "dar", "dice", "decir", "queremos", "quiero", "necesito", "necesita",
    # muletillas / saludos / genéricos
    "hola", "buenas", "buenos", "gracias", "saludos", "tardes", "dias", "noches", "favor",
    "ahora", "antes", "luego", "muy", "mas", "menos", "tan", "todo", "toda", "todos", "todas",
    "algo", "nada", "aun", "asi", "vez", "veces", "dia", "han", "hay", "the", "and",
    "tienda", "incidencia", "incidencias", "problema", "problemas", "cosa", "cosas",
}


def _tokens(texto: str) -> set:
    """Palabras significativas de un texto (minúsculas, sin acentos, >=3 letras)."""
    t = unicodedata.normalize("NFKD", (texto or "").lower())
    t = "".join(c for c in t if not unicodedata.combining(c))
    return {p for p in re.findall(r"[a-z0-9]{3,}", t) if p not in _STOPWORDS}


def _sugerir_faq(db, texto_incidencia: str, limite: int = 3) -> list:
    """Sugiere preguntas de la Biblioteca parecidas a la incidencia. Reglas:
       - Deben coincidir AL MENOS 2 palabras (no una suelta).
       - Alguna de esas palabras debe ser RELEVANTE (no salir en casi todas las
         preguntas): las palabras raras pesan más que las muy repetidas.
       - El título pesa el doble que el cuerpo."""
    from collections import Counter
    inc = _tokens(texto_incidencia)
    if not inc:
        return []
    faqs = db.query(models.conocimiento.Conocimiento).all()
    if not faqs:
        return []
    N = len(faqs)

    # Frecuencia de cada palabra entre las preguntas (para saber cuáles son comunes)
    df = Counter()
    datos = []
    for c in faqs:
        tt = _tokens(c.titulo)
        allt = tt | _tokens((c.contenido or "") + " " + (c.etiquetas or ""))
        datos.append((c, tt, allt))
        for w in allt:
            df[w] += 1

    # "relevante" = palabra que NO aparece en más de la mitad de las preguntas
    # (las que salen en casi todas son tan poco útiles como los artículos).
    limite_comun = max(1, N * 0.5)
    MIN_RELEVANTES = 2     # hacen falta al menos 2 palabras relevantes en común

    puntuadas = []
    for c, tt, allt in datos:
        comunes = inc & allt
        relevantes = {w for w in comunes if df[w] <= limite_comun}
        if len(relevantes) < MIN_RELEVANTES:
            continue       # si no coinciden bastantes palabras relevantes, no se sugiere
        # Las palabras raras valen más (1/frecuencia); el título cuenta doble
        score = sum((2.0 if w in tt else 1.0) / df[w] for w in relevantes)
        puntuadas.append((score, c))

    puntuadas.sort(key=lambda x: -x[0])
    return [c for _, c in puntuadas[:limite]]

# Aseguramos que se creen las tablas en tu archivo 'sql_app.db'
Base.metadata.create_all(bind=engine)


def _migrar_columnas():
    """Migración ligera para SQLite: añade columnas nuevas a tablas ya existentes."""
    from sqlalchemy import text
    with engine.begin() as conn:
        cols = {row[1] for row in conn.execute(text("PRAGMA table_info(tickets)"))}
        if "updated_at" not in cols:
            conn.execute(text("ALTER TABLE tickets ADD COLUMN updated_at DATETIME"))
            # Para los tickets antiguos, partimos de su fecha de creación
            conn.execute(text("UPDATE tickets SET updated_at = created_at WHERE updated_at IS NULL"))
        # Campos del flujo de trabajo (añadidos después)
        for col, tipo in (
            ("hora_llegada", "DATETIME"), ("hora_cierre", "DATETIME"), ("solucion", "TEXT"),
            ("nuevo_mensaje", "BOOLEAN DEFAULT 0"), ("email_inicial_id", "TEXT"),
            ("proximo_paso", "TEXT"), ("resumen", "TEXT DEFAULT ''"),
            ("cuerpo_inicial", "TEXT"),
        ):
            if col not in cols:
                conn.execute(text(f"ALTER TABLE tickets ADD COLUMN {col} {tipo}"))
        # La hora de llegada es la de llegada del correo: rellenamos las que falten
        conn.execute(text("UPDATE tickets SET hora_llegada = created_at WHERE hora_llegada IS NULL"))
        # El correo inicial del hilo parte de la descripción actual (para tickets ya existentes)
        conn.execute(text("UPDATE tickets SET cuerpo_inicial = descripcion WHERE cuerpo_inicial IS NULL"))
        # Columna nueva en mensajes
        mcols = {row[1] for row in conn.execute(text("PRAGMA table_info(messages)"))}
        if "outlook_message_id" not in mcols:
            conn.execute(text("ALTER TABLE messages ADD COLUMN outlook_message_id TEXT"))
        # Columnas nuevas en la biblioteca (conocimiento)
        ccols = {row[1] for row in conn.execute(text("PRAGMA table_info(conocimiento)"))}
        for col, tipo in (
            ("carpeta_id", "INTEGER"), ("tipo", "TEXT DEFAULT 'nota'"), ("url", "TEXT DEFAULT ''"),
            ("archivo_nombre", "TEXT DEFAULT ''"), ("archivo_ruta", "TEXT DEFAULT ''"),
            ("orden", "INTEGER DEFAULT 0"),
        ):
            if col not in ccols:
                conn.execute(text(f"ALTER TABLE conocimiento ADD COLUMN {col} {tipo}"))
                if col == "orden":  # orden inicial = id (orden de creación)
                    conn.execute(text("UPDATE conocimiento SET orden = id"))


_migrar_columnas()

app = FastAPI(title="Mi Gestor de Incidencias RFID")

app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")


@app.on_event("startup")
def arrancar_listener_outlook():
    """Lanza el vigilante de correos de Outlook en segundo plano al iniciar la web.
    Si Outlook no está configurado, la web sigue funcionando con normalidad.
    Esta función nunca debe lanzar excepciones: la web debe arrancar siempre."""
    try:
        import os
        import threading

        # Permite desactivarlo poniendo ACTIVAR_LISTENER=0 en el .env
        if os.getenv("ACTIVAR_LISTENER", "1") not in ("1", "true", "True"):
            print("[INFO] Listener de Outlook desactivado (ACTIVAR_LISTENER=0).")
            return

        from app.services.email_listener import start_email_listener

        def _runner():
            try:
                start_email_listener()
            except Exception as e:
                print(f"[AVISO] El listener de Outlook se detuvo: {e}")
                print("        La aplicación web sigue funcionando con normalidad.")

        hilo = threading.Thread(target=_runner, daemon=True, name="outlook-listener")
        hilo.start()
        print("[INFO] Listener de Outlook lanzado en segundo plano.")
    except Exception as e:
        print(f"[AVISO] No se pudo iniciar el listener de Outlook: {e}")

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# --- WEBHOOKS DE ENTRADA (Para conectar con tu script de Outlook) ---

class EmailPayload(BaseModel):
    body: str
    conversation_id: str
    sender: str

@app.post("/api/v1/webhooks/new-email")
def webhook_nuevo_correo(payload: EmailPayload, db: Session = Depends(get_db)):
    """Tu script de Outlook llamará aquí al detectar un correo con la plantilla"""
    ticket = TicketService.create_from_email(
        db=db, email_body=payload.body, conversation_id=payload.conversation_id
    )
    # Devolvemos el ID generado para que tu script sepa cómo renombrar el correo
    return {"status": "procesado", "ticket_id": ticket.ticket_id}

@app.post("/api/v1/webhooks/email-reply")
def webhook_respuesta_hilo(payload: EmailPayload, db: Session = Depends(get_db)):
    """Tu script de Outlook llamará aquí si alguien responde al hilo de un correo"""
    ticket = TicketService.append_reply_to_thread(
        db=db, conversation_id=payload.conversation_id, sender=payload.sender, body=payload.body
    )
    if not ticket:
        raise HTTPException(status_code=404, detail="Este hilo no pertenece a ningún ticket activo.")
    return {"status": "mensaje_añadido", "ticket_id": ticket.ticket_id}


# --- API JSON (consumida por el dashboard estilo Outlook) ---

def _fmt_tamano(b: int) -> str:
    b = b or 0
    if b < 1024:
        return f"{b} B"
    if b < 1024 * 1024:
        return f"{b/1024:.0f} KB"
    return f"{b/1024/1024:.1f} MB"


# Extensiones de vídeo que el navegador puede reproducir en línea
VIDEO_EXT = {".mp4", ".webm", ".mov", ".m4v", ".ogg", ".ogv"}


def _adjunto_dict(a) -> dict:
    # El tipo guardado solo distingue imagen/documento; los vídeos se guardaron como
    # "documento". Los reclasificamos aquí por su extensión para poder reproducirlos.
    ext = os.path.splitext(a.nombre or a.ruta or "")[1].lower()
    tipo = "video" if ext in VIDEO_EXT else a.tipo
    return {
        "id": a.id,
        "tipo": tipo,                           # imagen | video | documento
        "nombre": a.nombre,
        "tamano": _fmt_tamano(a.tamano),
        "src": f"/api/v1/adjuntos/{a.id}/archivo",
    }


@app.get("/api/v1/adjuntos/{adjunto_id}/archivo")
def api_servir_adjunto(adjunto_id: int, db: Session = Depends(get_db)):
    """Sirve un adjunto. Las imágenes en línea (para verlas); los documentos como descarga."""
    a = db.query(models.adjunto.Adjunto).filter(models.adjunto.Adjunto.id == adjunto_id).first()
    if not a or not a.ruta or not os.path.isfile(a.ruta):
        raise HTTPException(status_code=404, detail="Adjunto no encontrado.")
    ext = os.path.splitext(a.nombre or a.ruta or "")[1].lower()
    # Vídeos: en línea y con su tipo MIME correcto para que se reproduzcan
    # (FileResponse admite solicitudes de rango, así el vídeo se puede adelantar).
    if ext in VIDEO_EXT:
        import mimetypes
        mt = mimetypes.guess_type(a.nombre or a.ruta)[0] or "video/mp4"
        return FileResponse(a.ruta, media_type=mt)
    # filename -> el navegador descarga; sin filename -> lo muestra en línea (imágenes)
    if a.tipo == "documento":
        return FileResponse(a.ruta, filename=a.nombre or "documento")
    return FileResponse(a.ruta)


def _sla_ticket(t) -> tuple:
    """Devuelve (estado_sla, horas_sin_responder) para una incidencia.
    estado_sla: 'ok' | 'riesgo' | 'vencida'. Solo aplica a incidencias abiertas
    que aún no hemos respondido (en horario laboral)."""
    if t.estado == "CERRADO":
        return "ok", None
    if any(_es_nuestro(m.remitente) for m in t.messages):
        return "ok", None   # ya respondida
    ini = t.hora_llegada or t.created_at
    if not ini:
        return "ok", None
    horas = _tiempo_laboral(ini, datetime.now()).total_seconds() / 3600
    if horas > SLA_RESPUESTA_HORAS:
        return "vencida", horas
    if horas > SLA_RESPUESTA_HORAS * 0.6:
        return "riesgo", horas
    return "ok", horas


def _ticket_resumen(t: models.ticket.Ticket) -> dict:
    """Serializa un ticket para la lista (panel central)."""
    ultimo = t.messages[-1] if t.messages else None
    preview = (ultimo.contenido if ultimo else t.descripcion) or ""
    # Lo que se ve en la tarjeta: el RESUMEN a mano si lo hay; si no, un
    # extracto del correo original.
    if (t.resumen or "").strip():
        desc_corta = re.sub(r"\s+", " ", t.resumen.strip())[:140]
    else:
        desc_corta = re.sub(r"\s+", " ", (t.descripcion or "").strip().replace("\n", " "))[:120]
    sla, horas_sr = _sla_ticket(t)
    return {
        "sla": sla,
        "horas_sin_responder": round(horas_sr, 1) if horas_sr is not None else None,
        "ticket_id": t.ticket_id,
        "empresa": t.empresa,
        "localizacion": t.localizacion,
        "telefono": t.telefono,
        "estado": t.estado,
        "descripcion_corta": desc_corta,
        "preview": preview.strip().replace("\n", " ")[:160],
        "num_mensajes": len(t.messages),
        "nuevo_mensaje": bool(t.nuevo_mensaje),
        "proximo_paso": t.proximo_paso,
        "created_at": t.created_at.isoformat() if t.created_at else None,
        # Fecha del último cambio de estado (≈ fecha de cierre para las cerradas)
        "updated_at": t.updated_at.isoformat() if t.updated_at else (
            t.created_at.isoformat() if t.created_at else None),
        "ultimo_movimiento": (ultimo.fecha_envio if ultimo else t.created_at).isoformat()
        if (ultimo or t.created_at) else None,
    }


@app.get("/api/v1/tickets")
def api_listar_tickets(db: Session = Depends(get_db)):
    """Devuelve todos los tickets ordenados por actividad (el front filtra/busca)."""
    tickets = (
        db.query(models.ticket.Ticket)
        .order_by(models.ticket.Ticket.created_at.asc())  # orden de llegada
        .all()
    )
    return {"tickets": [_ticket_resumen(t) for t in tickets]}


@app.get("/api/v1/tickets/{ticket_id}")
def api_detalle_ticket(ticket_id: str, db: Session = Depends(get_db)):
    """Detalle completo de un ticket + historial de mensajes (panel de lectura)."""
    t = (
        db.query(models.ticket.Ticket)
        .filter(models.ticket.Ticket.ticket_id == ticket_id)
        .first()
    )
    if not t:
        raise HTTPException(status_code=404, detail="El ticket no existe.")
    # Al abrir la incidencia, el mensaje nuevo deja de ser novedad (se quita el azul)
    if t.nuevo_mensaje:
        t.nuevo_mensaje = False
        db.commit()

    # Adjuntos del ticket, agrupados por mensaje (None = correo inicial)
    adjuntos = db.query(models.adjunto.Adjunto).filter(
        models.adjunto.Adjunto.ticket_id_fk == t.id).all()
    por_mensaje = {}
    for a in adjuntos:
        por_mensaje.setdefault(a.message_id_fk, []).append(_adjunto_dict(a))

    from app.services.email_utils import cuerpo_para_hilo
    mensajes = []
    for m in t.messages:
        texto, reenviado_a = cuerpo_para_hilo(m.contenido, es_nuestro=_es_nuestro(m.remitente))
        mensajes.append({
            "remitente": m.remitente,
            "contenido": texto,
            "reenviado_a": reenviado_a,   # a quién se reenvió (solo en nuestros correos)
            "fecha_envio": m.fecha_envio.isoformat() if m.fecha_envio else None,
            "adjuntos": por_mensaje.get(m.id, []),
        })

    return {
        "ticket_id": t.ticket_id,
        "empresa": t.empresa,
        "localizacion": t.localizacion,
        "telefono": t.telefono,
        "resumen": t.resumen or "",
        "descripcion": t.descripcion,
        # El correo inicial que se ve en el hilo es INMUTABLE (no cambia al editar la descripción)
        "cuerpo_inicial": t.cuerpo_inicial or t.descripcion,
        "estado": t.estado,
        "created_at": t.created_at.isoformat() if t.created_at else None,
        "hora_llegada": t.hora_llegada.isoformat() if t.hora_llegada else None,
        "hora_cierre": t.hora_cierre.isoformat() if t.hora_cierre else None,
        "solucion": t.solucion,
        "proximo_paso": t.proximo_paso,
        "adjuntos_inicial": por_mensaje.get(None, []),
        "mensajes": mensajes,
    }


@app.get("/api/v1/tickets/{ticket_id}/sugerencias")
def api_sugerencias(ticket_id: str, db: Session = Depends(get_db)):
    """Sugiere preguntas de la Biblioteca parecidas a esta incidencia."""
    t = (
        db.query(models.ticket.Ticket)
        .filter(models.ticket.Ticket.ticket_id == ticket_id)
        .first()
    )
    if not t:
        raise HTTPException(status_code=404, detail="El ticket no existe.")
    # Nos basamos en el PRIMER correo (la descripción inicial = el que plantea el
    # problema). No usamos las respuestas del hilo, que añaden ruido.
    texto = (t.descripcion or "").strip()
    if not texto and t.messages:
        texto = t.messages[0].contenido or ""
    nombres_tema = {c.id: c.nombre for c in db.query(models.conocimiento.Carpeta).all()}
    faqs = _sugerir_faq(db, texto)
    return {
        "sugerencias": [
            {
                "id": c.id,
                "pregunta": c.titulo,
                "respuesta": c.contenido or "",
                "tema": nombres_tema.get(c.carpeta_id, ""),
            }
            for c in faqs
        ]
    }


# ===== RECORDATORIOS (incidencias manuales, no llegan por correo) =====
def _recordatorio_dict(r) -> dict:
    return {
        "id": r.id,
        "titulo": r.titulo,
        "detalle": r.detalle or "",
        "empresa": r.empresa or "",
        "prioridad": r.prioridad or "normal",
        "hecho": bool(r.hecho),
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }


class RecordatorioPayload(BaseModel):
    titulo: Optional[str] = None
    detalle: Optional[str] = None
    empresa: Optional[str] = None
    prioridad: Optional[str] = None
    hecho: Optional[bool] = None


@app.get("/api/v1/recordatorios")
def api_listar_recordatorios(db: Session = Depends(get_db)):
    # Pendientes primero, luego por fecha (los más nuevos arriba)
    rs = (
        db.query(models.recordatorio.Recordatorio)
        .order_by(models.recordatorio.Recordatorio.hecho.asc(),
                  models.recordatorio.Recordatorio.created_at.desc())
        .all()
    )
    return {"recordatorios": [_recordatorio_dict(r) for r in rs]}


@app.post("/api/v1/recordatorios")
def api_crear_recordatorio(payload: RecordatorioPayload, db: Session = Depends(get_db)):
    if not (payload.titulo and payload.titulo.strip()):
        raise HTTPException(status_code=400, detail="El recordatorio necesita un texto.")
    r = models.recordatorio.Recordatorio(
        titulo=payload.titulo.strip(),
        detalle=(payload.detalle or "").strip(),
        empresa=(payload.empresa or "").strip(),
        prioridad=(payload.prioridad or "normal"),
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return _recordatorio_dict(r)


@app.put("/api/v1/recordatorios/{rec_id}")
def api_editar_recordatorio(rec_id: int, payload: RecordatorioPayload, db: Session = Depends(get_db)):
    r = db.query(models.recordatorio.Recordatorio).filter(
        models.recordatorio.Recordatorio.id == rec_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="El recordatorio no existe.")
    if payload.titulo is not None:
        r.titulo = payload.titulo.strip()
    if payload.detalle is not None:
        r.detalle = payload.detalle.strip()
    if payload.empresa is not None:
        r.empresa = payload.empresa.strip()
    if payload.prioridad is not None:
        r.prioridad = payload.prioridad
    if payload.hecho is not None:
        r.hecho = payload.hecho
    db.commit()
    db.refresh(r)
    return _recordatorio_dict(r)


@app.delete("/api/v1/recordatorios/{rec_id}")
def api_borrar_recordatorio(rec_id: int, db: Session = Depends(get_db)):
    r = db.query(models.recordatorio.Recordatorio).filter(
        models.recordatorio.Recordatorio.id == rec_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="El recordatorio no existe.")
    db.delete(r)
    db.commit()
    return {"status": "eliminado", "id": rec_id}


def _parse_dt(valor):
    """Convierte una cadena ISO (incluida la de datetime-local '2026-06-19T14:30')
    en datetime. Cadena vacía o None -> None."""
    if not valor:
        return None
    try:
        return datetime.fromisoformat(valor)
    except ValueError:
        return None


class TicketUpdate(BaseModel):
    empresa: Optional[str] = None
    localizacion: Optional[str] = None
    telefono: Optional[str] = None
    resumen: Optional[str] = None
    descripcion: Optional[str] = None
    hora_llegada: Optional[str] = None
    hora_cierre: Optional[str] = None
    solucion: Optional[str] = None
    proximo_paso: Optional[str] = None
    estado: Optional[str] = None


@app.put("/api/v1/tickets/{ticket_id}")
def api_guardar_ticket(ticket_id: str, payload: TicketUpdate, db: Session = Depends(get_db)):
    """Guarda los datos editables de la incidencia y, opcionalmente, cambia su estado.
    Reglas del flujo:
      - Al pasar a EN_PROGRESO se fija la hora de llegada si no la había.
      - Al pasar a CERRADO hace falta una solución y se fija la hora de cierre si no la había."""
    t = (
        db.query(models.ticket.Ticket)
        .filter(models.ticket.Ticket.ticket_id == ticket_id)
        .first()
    )
    if not t:
        raise HTTPException(status_code=404, detail="El ticket no existe.")

    # Campos de texto (solo si vienen en la petición)
    for campo in ("empresa", "localizacion", "telefono", "resumen", "descripcion", "solucion", "proximo_paso"):
        valor = getattr(payload, campo)
        if valor is not None:
            setattr(t, campo, valor.strip() if isinstance(valor, str) else valor)

    # Fechas
    if payload.hora_llegada is not None:
        t.hora_llegada = _parse_dt(payload.hora_llegada)
    if payload.hora_cierre is not None:
        t.hora_cierre = _parse_dt(payload.hora_cierre)

    # Cambio de estado con reglas
    if payload.estado is not None:
        nuevo = payload.estado.upper()
        if nuevo not in ESTADOS_VALIDOS:
            raise HTTPException(status_code=400, detail=f"Estado no válido: {nuevo}")
        if nuevo == "CERRADO" and not (t.solucion and t.solucion.strip()):
            raise HTTPException(
                status_code=400,
                detail="Para cerrar la incidencia debes indicar una solución.",
            )
        # La hora de llegada es automática (la del correo); no se toca al avanzar.
        if nuevo == "CERRADO" and not t.hora_cierre:
            t.hora_cierre = datetime.now()
        t.estado = nuevo

    db.commit()
    db.refresh(t)
    return {"status": "guardado", "ticket_id": t.ticket_id, "estado": t.estado}


class MensajePayload(BaseModel):
    contenido: str
    remitente: Optional[str] = None


@app.post("/api/v1/tickets/{ticket_id}/mensajes")
def api_add_mensaje(ticket_id: str, payload: MensajePayload, db: Session = Depends(get_db)):
    """Añade una nota/respuesta al hilo de la incidencia (queda en el Excel)."""
    if not payload.contenido or not payload.contenido.strip():
        raise HTTPException(status_code=400, detail="La nota está vacía.")
    t = (
        db.query(models.ticket.Ticket)
        .filter(models.ticket.Ticket.ticket_id == ticket_id)
        .first()
    )
    if not t:
        raise HTTPException(status_code=404, detail="El ticket no existe.")
    msg = models.message.Message(
        ticket_id_fk=t.id,
        remitente=(payload.remitente or "rfid@stcretail.com").strip(),
        contenido=payload.contenido.strip(),
        fecha_envio=datetime.now(),
    )
    db.add(msg)
    db.commit()
    return {"status": "añadido", "ticket_id": t.ticket_id}


class EstadoPayload(BaseModel):
    estado: str


@app.patch("/api/v1/tickets/{ticket_id}/estado")
def api_cambiar_estado(ticket_id: str, payload: EstadoPayload, db: Session = Depends(get_db)):
    """Cambia solo el estado (atajo). Para el flujo completo usa PUT /tickets/{id}."""
    estado = payload.estado.upper()
    if estado not in ESTADOS_VALIDOS:
        raise HTTPException(status_code=400, detail=f"Estado no válido: {estado}")
    t = (
        db.query(models.ticket.Ticket)
        .filter(models.ticket.Ticket.ticket_id == ticket_id)
        .first()
    )
    if not t:
        raise HTTPException(status_code=404, detail="El ticket no existe.")
    if estado == "CERRADO" and not (t.solucion and t.solucion.strip()):
        raise HTTPException(status_code=400, detail="Para cerrar la incidencia debes indicar una solución.")
    if estado == "EN_PROGRESO" and not t.hora_llegada:
        t.hora_llegada = datetime.now()
    if estado == "CERRADO" and not t.hora_cierre:
        t.hora_cierre = datetime.now()
    t.estado = estado
    db.commit()
    return {"status": "actualizado", "ticket_id": t.ticket_id, "estado": t.estado}


# --- EXPORTACIÓN A EXCEL (con selección de columnas) ---

# Columnas disponibles para exportar: clave -> (etiqueta, cómo se calcula desde el ticket)
def _conversacion_texto(t) -> str:
    lineas = [f"[{t.created_at:%d/%m/%Y %H:%M}] Descripción inicial: {t.descripcion or ''}"] if t.created_at else []
    for m in t.messages:
        fecha = f"{m.fecha_envio:%d/%m/%Y %H:%M}" if m.fecha_envio else ""
        lineas.append(f"[{fecha}] {m.remitente or ''}: {m.contenido or ''}")
    return "\n".join(lineas)


def _fmt(dt):
    return dt.strftime("%d/%m/%Y %H:%M") if dt else ""


EXPORT_COLUMNS = {
    "ticket_id":    ("ID",                        lambda t: t.ticket_id),
    "created_at":   ("Fecha inicio",              lambda t: _fmt(t.created_at)),
    "empresa":      ("Empresa",                   lambda t: t.empresa or ""),
    "localizacion": ("Localización",              lambda t: t.localizacion or ""),
    "telefono":     ("Teléfono",                  lambda t: t.telefono or ""),
    "descripcion":  ("Descripción",               lambda t: t.descripcion or ""),
    "estado":       ("Estado",                    lambda t: t.estado or ""),
    "hora_llegada": ("Hora de llegada",           lambda t: _fmt(t.hora_llegada)),
    "hora_cierre":  ("Hora de cierre",            lambda t: _fmt(t.hora_cierre)),
    "solucion":     ("Solución",                  lambda t: t.solucion or ""),
    "proximo_paso": ("Próximo paso",              lambda t: t.proximo_paso or ""),
    "correo":       ("Correo",                    lambda t: (t.messages[0].remitente if t.messages else "")),
    "conversacion": ("Procedimiento/comentarios", _conversacion_texto),
    "num_mensajes": ("Nº mensajes",               lambda t: len(t.messages)),
    "updated_at":   ("Última actualización",      lambda t: _fmt(t.updated_at)),
}


# Horario laboral: L-V de 08:00 a 17:00 (se ignoran noches y fines de semana)
JORNADA_INI, JORNADA_FIN = 8, 17


def _tiempo_laboral(inicio, fin):
    """Tiempo transcurrido entre dos fechas contando SOLO horario laboral
    (lunes a viernes, de 08:00 a 17:00). Devuelve un timedelta."""
    if not inicio or not fin or fin <= inicio:
        return timedelta()
    total = timedelta()
    actual = inicio
    while actual < fin:
        dia = actual.date()
        if actual.weekday() < 5:  # 0-4 = lunes a viernes
            ini_j = datetime.combine(dia, time(JORNADA_INI, 0))
            fin_j = datetime.combine(dia, time(JORNADA_FIN, 0))
            tramo_ini = max(actual, ini_j)
            tramo_fin = min(fin, fin_j)
            if tramo_fin > tramo_ini:
                total += tramo_fin - tramo_ini
        actual = datetime.combine(dia + timedelta(days=1), time(0, 0))
    return total


def _fmt_duracion(td):
    """Formatea una duración (en horas laborables) como '13 h 30 min'."""
    if td is None:
        return "—"
    minutos = int(round(td.total_seconds() / 60))
    horas, mins = divmod(minutos, 60)
    if horas and mins:
        return f"{horas} h {mins} min"
    if horas:
        return f"{horas} h"
    return f"{mins} min"


def _calcular_kpis(tickets):
    """Calcula los indicadores (KPIs) sobre el conjunto de incidencias dado.
    Todos los tiempos se miden en horario laboral (L-V, 08:00-17:00)."""
    total = len(tickets)
    nuevas = sum(1 for t in tickets if t.estado == "PROCESADO")
    enprog = sum(1 for t in tickets if t.estado == "EN_PROGRESO")
    cerradas = [t for t in tickets if t.estado == "CERRADO"]
    pct = (len(cerradas) / total * 100) if total else 0

    # Tiempos de resolución (de la llegada al cierre, en horario laboral)
    duraciones = []
    for t in cerradas:
        ini = t.hora_llegada or t.created_at
        if ini and t.hora_cierre and t.hora_cierre >= ini:
            duraciones.append(_tiempo_laboral(ini, t.hora_cierre))
    media_res = (sum(duraciones, timedelta()) / len(duraciones)) if duraciones else None
    max_res = max(duraciones) if duraciones else None

    # Tiempo de respuesta: de la llegada a NUESTRA primera respuesta (horario laboral)
    import re as _re
    es_nuestro = lambda rem: bool(_re.search(r"sistema|software|stcretail", rem or "", _re.I))
    respuestas = []
    for t in tickets:
        ini = t.hora_llegada or t.created_at
        if ini:
            propios = [m.fecha_envio for m in t.messages if m.fecha_envio and es_nuestro(m.remitente)]
            if propios:
                primero = min(propios)
                if primero >= ini:
                    respuestas.append(_tiempo_laboral(ini, primero))
    media_resp = (sum(respuestas, timedelta()) / len(respuestas)) if respuestas else None

    # Antigüedad de las incidencias todavía abiertas (horario laboral)
    ahora = datetime.now()
    antig = []
    for t in tickets:
        if t.estado != "CERRADO":
            ini = t.hora_llegada or t.created_at
            if ini:
                antig.append(_tiempo_laboral(ini, ahora))
    media_abiertas = (sum(antig, timedelta()) / len(antig)) if antig else None
    # "Más de 3 días" => más de 3 jornadas laborales (3 x 9 h)
    mas_3d = sum(1 for d in antig if d.total_seconds() >= 3 * (JORNADA_FIN - JORNADA_INI) * 3600)

    filas = [
        ("Total de incidencias", total),
        ("Nuevas", nuevas),
        ("En progreso", enprog),
        ("Cerradas", len(cerradas)),
        ("% resueltas", f"{pct:.0f}%"),
        ("Tiempo medio de respuesta", _fmt_duracion(media_resp)),
        ("Tiempo medio de resolución", _fmt_duracion(media_res)),
        ("Tiempo máximo de resolución", _fmt_duracion(max_res)),
        ("Antigüedad media de las abiertas", _fmt_duracion(media_abiertas)),
        ("Abiertas con más de 3 días", mas_3d),
    ]
    datos = {"nuevas": nuevas, "enprog": enprog, "cerradas": len(cerradas), "abiertas": total - len(cerradas)}
    return filas, datos


@app.get("/api/v1/export/columns")
def api_export_columnas():
    """Lista las columnas disponibles para exportar (para el selector del front)."""
    return {"columnas": [{"key": k, "label": v[0]} for k, v in EXPORT_COLUMNS.items()]}


def _marcas_bd(db) -> list:
    """Empresas que aparecen en las incidencias guardadas (orden alfabético)."""
    filas = db.query(models.ticket.Ticket.empresa).distinct().all()
    return sorted({(f[0] or "").strip() for f in filas if f and f[0] and f[0].strip()})


def _lista_empresas(db) -> list:
    """La lista oficial de empresas: la de 'empresas.json'. Si ese archivo falta
    o está vacío, tiramos de las empresas que ya hay en las incidencias, para que
    los desplegables nunca se queden en blanco."""
    from app.services.empresas_service import cargar_empresas
    return cargar_empresas() or _marcas_bd(db)


@app.get("/api/v1/empresas")
def api_empresas(db: Session = Depends(get_db)):
    """Opciones del desplegable 'Empresa'. Se editan en 'empresas.json'."""
    return {"empresas": _lista_empresas(db)}


@app.get("/api/v1/export/marcas")
def api_export_marcas(db: Session = Depends(get_db)):
    """Lista de marcas/empresas para el selector del export."""
    return {"marcas": _lista_empresas(db)}


@app.get("/api/v1/export/excel")
def api_export_excel(
    columns: Optional[str] = None,
    estado: Optional[str] = None,
    empresa: Optional[str] = None,
    q: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Genera un Excel con las columnas elegidas, filtrando por estado y marca."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    # Columnas elegidas (en orden); si no se indican, todas
    pedidas = [c.strip() for c in columns.split(",")] if columns else list(EXPORT_COLUMNS)
    cols = [c for c in pedidas if c in EXPORT_COLUMNS] or list(EXPORT_COLUMNS)

    # Filtrado por estado y marca
    query = db.query(models.ticket.Ticket).order_by(models.ticket.Ticket.created_at.asc())
    if estado and estado not in ("TODOS", "TODAS"):
        query = query.filter(models.ticket.Ticket.estado == estado)
    if empresa and empresa not in ("TODAS", "TODOS"):
        query = query.filter(models.ticket.Ticket.empresa == empresa)
    tickets = query.all()
    if q:
        ql = q.lower()
        campos = ["empresa", "ticket_id", "localizacion", "telefono", "descripcion"]
        tickets = [t for t in tickets if any(ql in str(getattr(t, f) or "").lower() for f in campos)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidencias"

    # Cabecera con estilo
    ws.append([EXPORT_COLUMNS[c][0] for c in cols])
    azul = PatternFill("solid", fgColor="0F6CBD")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = azul
        cell.alignment = Alignment(vertical="center")

    # Filas
    for t in tickets:
        ws.append([EXPORT_COLUMNS[c][1](t) for c in cols])

    # Texto largo: se muestra completo (envuelto), no cortado a una sola línea
    largas = {"descripcion", "conversacion", "solucion"}
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    for fila in ws.iter_rows(min_row=2):
        for idx, celda in enumerate(fila):
            celda.alignment = wrap if cols[idx] in largas else top

    # Anchos de columna aproximados
    anchos = {"descripcion": 60, "conversacion": 70, "empresa": 22, "localizacion": 28, "correo": 30, "solucion": 45}
    for idx, c in enumerate(cols, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = anchos.get(c, 18)
    ws.freeze_panes = "A2"  # fija la cabecera

    # --- Hoja de RESUMEN / KPIs (segunda pestaña, tras los datos) ---
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint

    kpis, datos = _calcular_kpis(tickets)
    etiqueta_estado = {"PROCESADO": "Nuevas", "EN_PROGRESO": "En progreso", "CERRADO": "Cerradas"}
    resumen = wb.create_sheet("Resumen y KPIs")  # se añade después de "Incidencias"

    # Título + meta
    resumen["A1"] = "Resumen de incidencias"
    resumen["A1"].font = Font(bold=True, size=15, color="0F6CBD")
    meta = [("Generado:", datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("Marca:", empresa if (empresa and empresa not in ("TODAS", "TODOS")) else "Todas"),
            ("Estado:", etiqueta_estado.get(estado, "Todas"))]
    for i, (k, v) in enumerate(meta, start=3):
        resumen.cell(row=i, column=1, value=k).font = Font(bold=True, color="605E5C")
        resumen.cell(row=i, column=2, value=v)

    # Tabla de KPIs con bordes y cabecera
    thin = Side(style="thin", color="D0CECE")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    fila_h = 7
    cab = {1: "Indicador", 2: "Valor"}
    for col, txt in cab.items():
        c = resumen.cell(row=fila_h, column=col, value=txt)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F6CBD")
        c.border = borde
        c.alignment = Alignment(vertical="center", horizontal=("left" if col == 1 else "center"))
    r = fila_h + 1
    for i, (label, valor) in enumerate(kpis):
        cl = resumen.cell(row=r, column=1, value=label)
        cl.font = Font(bold=True)
        cl.border = borde
        cv = resumen.cell(row=r, column=2, value=valor)
        cv.border = borde
        cv.alignment = Alignment(horizontal="center")
        if i % 2 == 1:  # filas alternas sombreadas
            sombra = PatternFill("solid", fgColor="F2F6FB")
            cl.fill = sombra
            cv.fill = sombra
        r += 1

    resumen.column_dimensions["A"].width = 34
    resumen.column_dimensions["B"].width = 22

    # Datos para los gráficos en columnas P/Q (lejos a la derecha, fuera de la vista
    # de la tabla y de las gráficas; no se ocultan para que los gráficos las dibujen).
    def _set(rr, cc, vv):
        resumen.cell(row=rr, column=cc, value=vv)
    _set(1, 16, "Estado"); _set(1, 17, "Cantidad")
    _set(2, 16, "Nuevas"); _set(2, 17, datos["nuevas"])
    _set(3, 16, "En progreso"); _set(3, 17, datos["enprog"])
    _set(4, 16, "Cerradas"); _set(4, 17, datos["cerradas"])
    _set(6, 16, "Estado"); _set(6, 17, "Cantidad")
    _set(7, 16, "Completadas"); _set(7, 17, datos["cerradas"])
    _set(8, 16, "Sin completar"); _set(8, 17, datos["abiertas"])

    COLOR_NUEVAS, COLOR_PROG, COLOR_CERRADAS = "0F6CBD", "C19C00", "6B6B6B"

    # Gráfico de barras: incidencias por estado (solo el valor encima de cada barra)
    barras = BarChart()
    barras.title = "Incidencias por estado"
    barras.type = "col"
    barras.style = 10
    barras.legend = None
    barras.add_data(Reference(resumen, min_col=17, min_row=1, max_row=4), titles_from_data=True)
    barras.set_categories(Reference(resumen, min_col=16, min_row=2, max_row=4))
    dl = DataLabelList()
    dl.showVal, dl.showSerName, dl.showCatName, dl.showLegendKey, dl.showPercent = True, False, False, False, False
    barras.dataLabels = dl
    serie = barras.series[0]
    serie.graphicalProperties.line.noFill = True
    for i, color in enumerate((COLOR_NUEVAS, COLOR_PROG, COLOR_CERRADAS)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        serie.data_points.append(pt)
    barras.height, barras.width = 8, 14
    resumen.add_chart(barras, "D2")

    # Gráfico circular: completadas vs sin completar (solo el porcentaje)
    tarta = PieChart()
    tarta.title = "Completadas vs sin completar"
    tarta.style = 10
    tarta.add_data(Reference(resumen, min_col=17, min_row=6, max_row=8), titles_from_data=True)
    tarta.set_categories(Reference(resumen, min_col=16, min_row=7, max_row=8))
    dl2 = DataLabelList()
    dl2.showPercent, dl2.showVal, dl2.showSerName, dl2.showCatName, dl2.showLegendKey = True, False, False, False, False
    tarta.dataLabels = dl2
    s2 = tarta.series[0]
    for i, color in enumerate((COLOR_CERRADAS, COLOR_NUEVAS)):  # Completadas, Sin completar
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        s2.data_points.append(pt)
    tarta.height, tarta.width = 8, 14
    resumen.add_chart(tarta, "D21")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"incidencias_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# --- BIBLIOTECA (carpetas + items: notas, enlaces y archivos) ---

BIBLIOTECA_DIR = "biblioteca_files"
os.makedirs(BIBLIOTECA_DIR, exist_ok=True)


def _media_dict(m) -> dict:
    return {
        "id": m.id,
        "tipo": m.tipo,                          # imagen | video | enlace_video
        "nombre": m.archivo_nombre or "",
        "url": m.url or "",                       # para enlace_video (YouTube…)
        "src": (f"/api/v1/media/{m.id}/archivo" if m.archivo_ruta else (m.url or "")),
    }


def _item_dict(c, con_media: bool = False) -> dict:
    d = {
        "id": c.id,
        "carpeta_id": c.carpeta_id,
        "titulo": c.titulo,                       # la pregunta
        "etiquetas": c.etiquetas or "",
        "contenido": c.contenido or "",           # la respuesta
        "num_media": len(c.media),
        "updated_at": c.updated_at.isoformat() if c.updated_at else None,
    }
    if con_media:
        d["media"] = [_media_dict(m) for m in c.media]
    return d


# ----- Carpetas -----
class CarpetaPayload(BaseModel):
    nombre: str


@app.get("/api/v1/carpetas")
def api_listar_carpetas(db: Session = Depends(get_db)):
    carpetas = db.query(models.conocimiento.Carpeta).order_by(models.conocimiento.Carpeta.nombre).all()
    # contar items por carpeta
    conteo = {}
    for it in db.query(models.conocimiento.Conocimiento).all():
        conteo[it.carpeta_id] = conteo.get(it.carpeta_id, 0) + 1
    return {
        "carpetas": [{"id": c.id, "nombre": c.nombre, "num": conteo.get(c.id, 0)} for c in carpetas],
        "sin_carpeta": conteo.get(None, 0),
        "total": sum(conteo.values()),
    }


@app.post("/api/v1/carpetas")
def api_crear_carpeta(payload: CarpetaPayload, db: Session = Depends(get_db)):
    if not payload.nombre.strip():
        raise HTTPException(status_code=400, detail="El nombre es obligatorio.")
    c = models.conocimiento.Carpeta(nombre=payload.nombre.strip())
    db.add(c)
    db.commit()
    db.refresh(c)
    return {"id": c.id, "nombre": c.nombre, "num": 0}


@app.put("/api/v1/carpetas/{carpeta_id}")
def api_renombrar_carpeta(carpeta_id: int, payload: CarpetaPayload, db: Session = Depends(get_db)):
    c = db.query(models.conocimiento.Carpeta).filter(models.conocimiento.Carpeta.id == carpeta_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="La carpeta no existe.")
    if not payload.nombre.strip():
        raise HTTPException(status_code=400, detail="El nombre es obligatorio.")
    c.nombre = payload.nombre.strip()
    db.commit()
    return {"id": c.id, "nombre": c.nombre}


@app.delete("/api/v1/carpetas/{carpeta_id}")
def api_borrar_carpeta(carpeta_id: int, db: Session = Depends(get_db)):
    c = db.query(models.conocimiento.Carpeta).filter(models.conocimiento.Carpeta.id == carpeta_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="La carpeta no existe.")
    # Los items de la carpeta pasan a "sin carpeta" (no se borran)
    for it in db.query(models.conocimiento.Conocimiento).filter(
            models.conocimiento.Conocimiento.carpeta_id == carpeta_id).all():
        it.carpeta_id = None
    db.delete(c)
    db.commit()
    return {"status": "eliminada", "id": carpeta_id}


# ----- Preguntas (items) -----
class ItemPayload(BaseModel):
    titulo: Optional[str] = None         # la pregunta
    carpeta_id: Optional[int] = None     # el tema
    etiquetas: Optional[str] = None
    contenido: Optional[str] = None      # la respuesta


@app.get("/api/v1/biblioteca")
def api_listar_biblioteca(carpeta: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(models.conocimiento.Conocimiento)
    if carpeta is not None:
        q = q.filter(models.conocimiento.Conocimiento.carpeta_id == carpeta)
    items = q.order_by(
        models.conocimiento.Conocimiento.orden.asc(),
        models.conocimiento.Conocimiento.id.asc(),
    ).all()
    return {"items": [_item_dict(c) for c in items]}


class OrdenPayload(BaseModel):
    ids: list[int]


@app.post("/api/v1/biblioteca/orden")
def api_guardar_orden(payload: OrdenPayload, db: Session = Depends(get_db)):
    """Guarda el orden manual de las preguntas (lista de ids en el orden deseado)."""
    for posicion, item_id in enumerate(payload.ids):
        c = db.query(models.conocimiento.Conocimiento).filter(
            models.conocimiento.Conocimiento.id == item_id).first()
        if c:
            c.orden = posicion
    db.commit()
    return {"status": "ok"}


@app.get("/api/v1/biblioteca/{item_id}")
def api_detalle_item(item_id: int, db: Session = Depends(get_db)):
    c = db.query(models.conocimiento.Conocimiento).filter(
        models.conocimiento.Conocimiento.id == item_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="La pregunta no existe.")
    return _item_dict(c, con_media=True)


@app.post("/api/v1/biblioteca")
def api_crear_item(payload: ItemPayload, db: Session = Depends(get_db)):
    if not (payload.titulo and payload.titulo.strip()):
        raise HTTPException(status_code=400, detail="La pregunta es obligatoria.")
    from sqlalchemy import func
    max_orden = db.query(func.max(models.conocimiento.Conocimiento.orden)).scalar() or 0
    c = models.conocimiento.Conocimiento(
        titulo=payload.titulo.strip(),
        carpeta_id=payload.carpeta_id or None,
        etiquetas=(payload.etiquetas or "").strip(),
        contenido=(payload.contenido or "").strip(),
        orden=max_orden + 1,
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return _item_dict(c, con_media=True)


@app.put("/api/v1/biblioteca/{item_id}")
def api_guardar_item(item_id: int, payload: ItemPayload, db: Session = Depends(get_db)):
    c = db.query(models.conocimiento.Conocimiento).filter(
        models.conocimiento.Conocimiento.id == item_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="La pregunta no existe.")
    if payload.titulo is not None:
        if not payload.titulo.strip():
            raise HTTPException(status_code=400, detail="La pregunta es obligatoria.")
        c.titulo = payload.titulo.strip()
    if payload.etiquetas is not None:
        c.etiquetas = payload.etiquetas.strip()
    if payload.contenido is not None:
        c.contenido = payload.contenido.strip()
    if payload.carpeta_id is not None:
        c.carpeta_id = payload.carpeta_id or None
    db.commit()
    db.refresh(c)
    return _item_dict(c, con_media=True)


@app.delete("/api/v1/biblioteca/{item_id}")
def api_borrar_item(item_id: int, db: Session = Depends(get_db)):
    c = db.query(models.conocimiento.Conocimiento).filter(
        models.conocimiento.Conocimiento.id == item_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="La pregunta no existe.")
    # Borra los archivos físicos de su media
    for m in c.media:
        if m.archivo_ruta and os.path.isfile(m.archivo_ruta):
            try:
                os.remove(m.archivo_ruta)
            except OSError:
                pass
    db.delete(c)
    db.commit()
    return {"status": "eliminado", "id": item_id}


# ----- Media (fotos / vídeos de una pregunta) -----
IMG_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
VID_EXT = {".mp4", ".webm", ".mov", ".m4v", ".avi", ".ogg"}


@app.post("/api/v1/biblioteca/{item_id}/media")
async def api_subir_media(item_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Sube una foto o vídeo y lo adjunta a la pregunta."""
    c = db.query(models.conocimiento.Conocimiento).filter(
        models.conocimiento.Conocimiento.id == item_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="La pregunta no existe.")
    nombre = os.path.basename(file.filename or "archivo")
    ext = os.path.splitext(nombre)[1].lower()
    tipo = "imagen" if ext in IMG_EXT else "video" if ext in VID_EXT else "documento"
    ruta = os.path.join(BIBLIOTECA_DIR, f"{uuid.uuid4().hex}{ext}")
    with open(ruta, "wb") as f:
        f.write(await file.read())
    m = models.conocimiento.Media(item_id=item_id, tipo=tipo, archivo_nombre=nombre, archivo_ruta=ruta)
    db.add(m)
    db.commit()
    db.refresh(m)
    return _media_dict(m)


class MediaURLPayload(BaseModel):
    url: str


@app.post("/api/v1/biblioteca/{item_id}/media/url")
def api_media_url(item_id: int, payload: MediaURLPayload, db: Session = Depends(get_db)):
    """Adjunta un vídeo por enlace (YouTube, etc.) a la pregunta."""
    c = db.query(models.conocimiento.Conocimiento).filter(
        models.conocimiento.Conocimiento.id == item_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="La pregunta no existe.")
    if not payload.url.strip():
        raise HTTPException(status_code=400, detail="El enlace está vacío.")
    m = models.conocimiento.Media(item_id=item_id, tipo="enlace_video", url=payload.url.strip())
    db.add(m)
    db.commit()
    db.refresh(m)
    return _media_dict(m)


@app.delete("/api/v1/media/{media_id}")
def api_borrar_media(media_id: int, db: Session = Depends(get_db)):
    m = db.query(models.conocimiento.Media).filter(models.conocimiento.Media.id == media_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="El archivo no existe.")
    if m.archivo_ruta and os.path.isfile(m.archivo_ruta):
        try:
            os.remove(m.archivo_ruta)
        except OSError:
            pass
    db.delete(m)
    db.commit()
    return {"status": "eliminado", "id": media_id}


@app.get("/api/v1/media/{media_id}/archivo")
def api_servir_media(media_id: int, db: Session = Depends(get_db)):
    m = db.query(models.conocimiento.Media).filter(models.conocimiento.Media.id == media_id).first()
    if not m or not m.archivo_ruta or not os.path.isfile(m.archivo_ruta):
        raise HTTPException(status_code=404, detail="Archivo no encontrado.")
    return FileResponse(m.archivo_ruta, filename=m.archivo_nombre or "archivo")


# --- PANEL DE KPIs (en vivo, dentro de la app) ---

@app.get("/api/v1/kpis")
def api_kpis(periodo: Optional[str] = None, marca: Optional[str] = None, db: Session = Depends(get_db)):
    from collections import Counter
    tickets = db.query(models.ticket.Ticket).all()

    # Filtro por marca
    if marca and marca not in ("TODAS", "TODOS"):
        tickets = [t for t in tickets if (t.empresa or "").strip() == marca]

    # Filtro por periodo (según la fecha de llegada de la incidencia)
    dias = {"semana": 7, "mes": 30, "anio": 365}.get(periodo)
    if dias:
        limite = datetime.now() - timedelta(days=dias)
        tickets = [t for t in tickets if (t.hora_llegada or t.created_at) and (t.hora_llegada or t.created_at) >= limite]

    filas, datos = _calcular_kpis(tickets)

    cnt_marca = Counter((t.empresa or "—").strip() or "—" for t in tickets)
    cnt_tienda = Counter((t.localizacion or "—").strip() or "—" for t in tickets)

    # SLA: incidencias abiertas sin responder que han superado el objetivo
    venc = riesgo = 0
    for t in tickets:
        s, _ = _sla_ticket(t)
        if s == "vencida":
            venc += 1
        elif s == "riesgo":
            riesgo += 1

    # Opciones del desplegable de marca: la lista oficial de 'empresas.json'
    marcas_disponibles = _lista_empresas(db)

    return {
        "kpis": [{"label": l, "valor": v} for l, v in filas],
        "datos": datos,
        "ranking_marca": cnt_marca.most_common(8),
        "ranking_tienda": cnt_tienda.most_common(8),
        "sla_vencidas": venc,
        "sla_riesgo": riesgo,
        "sla_horas": SLA_RESPUESTA_HORAS,
        "marcas": marcas_disponibles,
        "periodo": periodo or "siempre",
        "marca": marca or "TODAS",
    }


# --- VISTAS HTML DE TU PROPIO SISTEMA ---

@app.get("/biblioteca", response_class=HTMLResponse)
def ver_biblioteca(request: Request):
    """Página de la Biblioteca."""
    return templates.TemplateResponse(request, "biblioteca.html", {})


@app.get("/kpis", response_class=HTMLResponse)
def ver_kpis(request: Request):
    """Panel de KPIs en vivo."""
    return templates.TemplateResponse(request, "kpis.html", {})


@app.get("/", response_class=HTMLResponse)
def ver_dashboard(request: Request, db: Session = Depends(get_db)):
    """Tu panel de control principal: Lista todas las incidencias creadas"""
    tickets = db.query(models.ticket.Ticket).order_by(models.ticket.Ticket.created_at.desc()).all()
    return templates.TemplateResponse(request, "dashboard.html", {"tickets": tickets})

@app.get("/ticket/{ticket_id}", response_class=HTMLResponse)
def ver_detalle_ticket(ticket_id: str, request: Request, db: Session = Depends(get_db)):
    """El visor estilo conversación de un ticket concreto"""
    ticket = db.query(models.ticket.Ticket).filter(models.ticket.Ticket.ticket_id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="El ticket no existe.")
    return templates.TemplateResponse(request, "ticket.html", {
        "ticket": ticket,
        "messages": ticket.messages,
    })